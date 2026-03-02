#!/usr/bin/env python3
import argparse
import datetime as dt
import io
import mailbox
import os
import re
from email.utils import parsedate_to_datetime
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook

from orders_analytics.utils.constants import raw_path, takeout_path


def _clean_value(value: object) -> str:
    if value is None or pd.isna(value):
        return ""
    return str(value).strip()


def _find_order_id_column(columns: List[str]) -> Optional[str]:
    for column in columns:
        lowered = str(column).strip().lower().replace("_", " ")
        lowered = " ".join(lowered.split())
        if lowered in {"order id", "orderid"}:
            return column
    return None


ORDER_ID_FORMULA_RE = re.compile(r'^\s*=\s*"([^"]+)"\s*$')


def _normalize_date_text(value: str) -> str:
    text = _clean_value(value)
    if not text:
        return ""
    parsed = pd.to_datetime(text, errors="coerce")
    if pd.isna(parsed):
        return text
    return parsed.date().isoformat()


def _normalize_amount_text(value: str) -> str:
    text = _clean_value(value).replace("$", "").replace(",", "")
    if not text:
        return ""
    try:
        return f"{float(text):.2f}"
    except ValueError:
        return text


def _rows_from_dataframe(
    df: pd.DataFrame,
    provider: str,
    email_date: str,
    attachment_name: str,
    member_name: str,
    sheet_name: str,
) -> List[Dict[str, str]]:
    if df.empty:
        return []
    data = df.fillna("")
    data.columns = [str(col) for col in data.columns]
    order_col = _find_order_id_column(list(data.columns))
    if not order_col:
        return []

    rows: List[Dict[str, str]] = []
    for _, series in data.iterrows():
        row = {str(col): _clean_value(series.get(col, "")) for col in data.columns}
        order_id = _clean_value(row.get(order_col, ""))
        if not order_id:
            continue
        row["provider"] = provider
        row["source_file"] = attachment_name
        row["source_member"] = member_name or attachment_name
        row["source_sheet"] = sheet_name
        row["email_date"] = email_date
        rows.append(row)
    return rows


def _parse_csv_bytes(
    payload: bytes,
    provider: str,
    email_date: str,
    attachment_name: str,
    member_name: str,
) -> List[Dict[str, str]]:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = payload.decode(encoding, errors="strict")
            break
        except UnicodeDecodeError:
            continue
    else:
        text = payload.decode("utf-8", errors="ignore")
    df = pd.read_csv(io.StringIO(text), dtype=str)
    df = _normalize_csv_columns(df)
    return _rows_from_dataframe(df, provider, email_date, attachment_name, member_name, "")


def _to_amount(value: object) -> float:
    text = _clean_value(value).replace("$", "").replace(",", "")
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def _format_amount(value: float) -> str:
    return f"{value:.2f}"


def _normalize_csv_columns(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out.columns = [str(col) for col in out.columns]

    def fill_from(target: str, source: str) -> None:
        if source not in out.columns:
            return
        if target not in out.columns:
            out[target] = out[source]
            return
        source_values = out[source].fillna("").astype(str).str.strip()
        target_values = out[target].fillna("").astype(str).str.strip()
        out[target] = target_values.mask(target_values == "", source_values)

    fill_from("Date (Reportable)", "Date")
    fill_from("Order ID", "OrderId")
    if "OrderType" in out.columns:
        if "Transaction Type" not in out.columns:
            out["Transaction Type"] = ""
        txn_values = out["Transaction Type"].fillna("").astype(str).str.strip()
        out["Transaction Type"] = txn_values.mask(txn_values == "", "delivery")

    if "Subtotal" not in out.columns:
        out["Subtotal"] = ""
    subtotal_values = out["Subtotal"].fillna("").astype(str).str.strip()
    if "DeliverySubtotal" in out.columns or "PickupSubtotal" in out.columns:
        delivery_vals = (
            out["DeliverySubtotal"].astype(str).map(_to_amount)
            if "DeliverySubtotal" in out.columns
            else 0.0
        )
        pickup_vals = (
            out["PickupSubtotal"].astype(str).map(_to_amount)
            if "PickupSubtotal" in out.columns
            else 0.0
        )
        computed_subtotal = (delivery_vals + pickup_vals).map(_format_amount)
        out["Subtotal"] = subtotal_values.mask(subtotal_values == "", computed_subtotal)

    fill_from("Gross Tax", "GrossTax")
    fill_from("Tax Remitted", "TaxRemitted")
    fill_from("Total Promotions", "Savings")
    fill_from("Total Adjustments", "Adjustments")
    fill_from("Total Adjustments", "TotalAdjustments")

    # Explicit store-name correction requested for the lone CSV order.
    if "Order ID" in out.columns:
        order_ids = out["Order ID"].fillna("").astype(str).str.strip().str.lower()
        if "Store Name" not in out.columns:
            out["Store Name"] = ""
        out.loc[
            order_ids == "f7bd2429",
            "Store Name",
        ] = "Firebelly Wings 25431 Trabuco Road"

    # Drop CSV-only source columns after mapping so raw schema aligns with XLSX rows.
    drop_columns = [
        "Date",
        "OrderId",
        "OrderType",
        "DeliverySubtotal",
        "PickupSubtotal",
        "GrossTax",
        "TaxRemitted",
        "Savings",
        "Adjustments",
        "TotalAdjustments",
        "TaxPassthrough",
    ]
    existing_drop = [col for col in drop_columns if col in out.columns]
    if existing_drop:
        out = out.drop(columns=existing_drop)

    return out


def _parse_xlsx_bytes(
    payload: bytes,
    provider: str,
    email_date: str,
    attachment_name: str,
    member_name: str,
) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    try:
        sheets = pd.read_excel(io.BytesIO(payload), sheet_name=None, dtype=str, engine="openpyxl")
    except Exception:
        return []
    wb_formula = None
    try:
        wb_formula = load_workbook(io.BytesIO(payload), data_only=False)
    except Exception:
        wb_formula = None
    for sheet_name, df in sheets.items():
        # Recover string Order ID values stored as formulas (e.g. ="51ff14f8"),
        # which pandas may otherwise read as cached numeric 0.
        if wb_formula is not None and sheet_name in wb_formula.sheetnames:
            order_col = _find_order_id_column([str(c) for c in df.columns])
            if order_col:
                ws = wb_formula[sheet_name]
                col_idx = list(df.columns).index(order_col) + 1
                for row_idx in range(len(df)):
                    excel_row = row_idx + 2  # header in row 1
                    cell = ws.cell(excel_row, col_idx)
                    value = cell.value
                    if isinstance(value, str):
                        m = ORDER_ID_FORMULA_RE.match(value)
                        if m:
                            df.at[row_idx, order_col] = m.group(1)
        rows.extend(
            _rows_from_dataframe(
                df=df,
                provider=provider,
                email_date=email_date,
                attachment_name=attachment_name,
                member_name=member_name,
                sheet_name=str(sheet_name),
            )
        )
    return rows


def _normalize_header_name(value: str) -> str:
    text = _clean_value(value).lower()
    text = text.replace("\n", " ").replace("_", " ")
    text = " ".join(text.split())
    return text


def _extract_summary_rows_from_xlsx(
    payload: bytes,
    provider: str,
    email_date: str,
    attachment_name: str,
    member_name: str,
) -> List[Dict[str, str]]:
    try:
        df = pd.read_excel(io.BytesIO(payload), sheet_name="Summary", dtype=str, engine="openpyxl").fillna("")
    except Exception:
        return []
    if df.empty:
        return []

    header_idx = None
    headers: List[str] = []
    for idx, row in df.iterrows():
        values = [_clean_value(v) for v in row.tolist()]
        normalized = [_normalize_header_name(v) for v in values]
        if "pay period ending" in normalized and "dsp" in normalized:
            header_idx = idx
            headers = values
            break
    if header_idx is None:
        return []

    rows: List[Dict[str, str]] = []
    current_pay_period = ""
    current_store_name = ""
    current_location = ""

    for i in range(header_idx + 1, len(df)):
        raw_values = [_clean_value(v) for v in df.iloc[i].tolist()]
        if not any(raw_values):
            continue
        record = {}
        for col_idx, header in enumerate(headers):
            key = header if header else f"column_{col_idx + 1}"
            record[key] = raw_values[col_idx] if col_idx < len(raw_values) else ""

        pay_period = _clean_value(record.get("Pay Period Ending", ""))
        store_name = _clean_value(record.get("Store Name", ""))
        location = _clean_value(record.get("Location", ""))
        dsp = _clean_value(record.get("DSP", ""))
        concept = _clean_value(record.get("Concept", ""))

        if pay_period:
            current_pay_period = pay_period
        if store_name:
            current_store_name = store_name
        if location:
            current_location = location

        record["Pay Period Ending"] = current_pay_period
        record["Store Name"] = current_store_name
        record["Location"] = current_location

        if not dsp:
            continue
        if concept.lower() in {"total", "grand total"}:
            continue

        record["provider"] = provider
        record["source_file"] = attachment_name
        record["source_member"] = member_name or attachment_name
        record["source_sheet"] = dsp
        record["summary_sheet"] = "Summary"
        record["email_date"] = email_date
        rows.append(record)
    return rows


def _parse_payload(
    payload: bytes,
    filename: str,
    provider: str,
    email_date: str,
    member_name: str = "",
) -> List[Dict[str, str]]:
    lower = filename.lower()
    if lower.endswith(".xlsx"):
        return _parse_xlsx_bytes(payload, provider, email_date, filename, member_name)
    if lower.endswith(".csv"):
        return _parse_csv_bytes(payload, provider, email_date, filename, member_name)
    if lower.endswith(".zip"):
        # User-confirmed duplicate payloads; skip zip attachments for Nextbite.
        return []
    return []


def parse_mbox(mbox_path: str) -> Tuple[List[Dict[str, str]], List[Dict[str, str]]]:
    order_rows: List[Dict[str, str]] = []
    summary_rows: List[Dict[str, str]] = []
    mbox = mailbox.mbox(mbox_path)
    for msg in mbox:
        email_date = ""
        if msg.get("Date"):
            try:
                email_date = parsedate_to_datetime(msg.get("Date")).isoformat()
            except (TypeError, ValueError):
                email_date = ""
        if not msg.is_multipart():
            continue
        for part in msg.walk():
            filename = part.get_filename()
            if not filename:
                continue
            payload = part.get_payload(decode=True)
            if not payload:
                continue
            if filename.lower().endswith(".xlsx"):
                summary_rows.extend(
                    _extract_summary_rows_from_xlsx(
                        payload=payload,
                        provider="NEXTBITE",
                        email_date=email_date,
                        attachment_name=filename,
                        member_name=filename,
                    )
                )
            order_rows.extend(
                _parse_payload(
                    payload=payload,
                    filename=filename,
                    provider="NEXTBITE",
                    email_date=email_date,
                )
            )
    return order_rows, summary_rows


def dedupe_rows(rows: List[Dict[str, str]]) -> List[Dict[str, str]]:
    def key(row: Dict[str, str]) -> str:
        return "||".join(
            [
                _clean_value(row.get("Order ID", "")),
                _normalize_date_text(_clean_value(row.get("Date (Reportable)", ""))),
                _normalize_amount_text(_clean_value(row.get("Subtotal", ""))),
                _normalize_amount_text(_clean_value(row.get("Gross Tax", ""))),
                _normalize_amount_text(_clean_value(row.get("Tax Remitted", ""))),
                _normalize_amount_text(_clean_value(row.get("Total Promotions", ""))),
                _normalize_amount_text(_clean_value(row.get("Total Adjustments", ""))),
            ]
        )

    def score(row: Dict[str, str]) -> int:
        source_file = _clean_value(row.get("source_file", "")).lower()
        source_member = _clean_value(row.get("source_member", "")).lower()
        # Prefer workbook/email rows over CSV rows when duplicate keys collide.
        source_bonus = 10
        if source_file.endswith(".csv") or source_member.endswith(".csv"):
            source_bonus = 0
        return source_bonus + sum(1 for v in row.values() if _clean_value(v))

    chosen: Dict[str, Dict[str, str]] = {}
    for row in rows:
        k = key(row)
        current = chosen.get(k)
        if current is None:
            chosen[k] = row
            continue
        if score(row) > score(current):
            chosen[k] = row
    deduped = list(chosen.values())

    # Remove "Order ID = 0" duplicates when an equivalent non-zero-id row exists.
    def eq_key_without_order_id(row: Dict[str, str]) -> str:
        return "||".join(
            [
                _clean_value(row.get("Store Name", "")),
                _normalize_date_text(_clean_value(row.get("Date (Reportable)", ""))),
                _clean_value(row.get("source_sheet", "")).lower(),
                _clean_value(row.get("Transaction Type", "")).lower(),
                _normalize_amount_text(_clean_value(row.get("Subtotal", ""))),
                _normalize_amount_text(_clean_value(row.get("Gross Tax", ""))),
                _normalize_amount_text(_clean_value(row.get("Tax Remitted", ""))),
                _normalize_amount_text(_clean_value(row.get("Total Commissions", ""))),
                _normalize_amount_text(_clean_value(row.get("Total Promotions", ""))),
                _normalize_amount_text(_clean_value(row.get("Total Adjustments", ""))),
            ]
        )

    nonzero_keys = {
        eq_key_without_order_id(row)
        for row in deduped
        if _clean_value(row.get("Order ID", "")) not in {"", "0"}
    }
    filtered: List[Dict[str, str]] = []
    for row in deduped:
        order_id = _clean_value(row.get("Order ID", ""))
        if order_id == "0" and eq_key_without_order_id(row) in nonzero_keys:
            continue
        filtered.append(row)
    return filtered


def dedupe_summary_rows(rows: List[Dict[str, str]]) -> List[Dict[str, str]]:
    def key(row: Dict[str, str]) -> str:
        return "||".join(
            [
                _clean_value(row.get("Pay Period Ending", "")),
                _clean_value(row.get("Store Name", "")),
                _clean_value(row.get("Location", "")),
                _clean_value(row.get("DSP", "")).lower(),
                _clean_value(row.get("Concept", "")),
                _normalize_amount_text(_clean_value(row.get("Total Gross Sales", ""))),
                _normalize_amount_text(_clean_value(row.get("Total FP Payout", ""))),
            ]
        )

    def score(row: Dict[str, str]) -> int:
        return sum(1 for v in row.values() if _clean_value(v))

    chosen: Dict[str, Dict[str, str]] = {}
    for row in rows:
        k = key(row)
        current = chosen.get(k)
        if current is None or score(row) > score(current):
            chosen[k] = row
    return list(chosen.values())


def write_raw(path: str, rows: List[Dict[str, str]]) -> int:
    now = dt.datetime.now().isoformat()
    for row in rows:
        row.setdefault("added_at", now)

    all_columns: List[str] = []
    seen = set()
    for row in rows:
        for key in row.keys():
            if key not in seen:
                seen.add(key)
                all_columns.append(key)
    if "added_at" not in seen:
        all_columns.append("added_at")

    if "Commission %" in all_columns and "Savings" in all_columns:
        all_columns = [col for col in all_columns if col != "Savings"]
        insert_at = all_columns.index("Commission %") + 1
        all_columns.insert(insert_at, "Savings")

    # Keep parser-added metadata fields at the end so provider columns stay together.
    metadata_columns = [
        "provider",
        "source_file",
        "source_member",
        "source_sheet",
        "summary_sheet",
        "email_date",
        "added_at",
    ]
    all_columns = [col for col in all_columns if col not in metadata_columns] + [
        col for col in metadata_columns if col in all_columns
    ]

    os.makedirs(os.path.dirname(path), exist_ok=True)
    pd.DataFrame(rows).reindex(columns=all_columns).to_csv(path, index=False)
    return len(rows)


def run(
    mbox_path: str,
    out_path: str = raw_path("nextbite", "orders_raw.csv"),
    summary_out_path: str = raw_path("nextbite", "billings_raw.csv"),
) -> int:
    order_rows, summary_rows = parse_mbox(mbox_path)
    order_rows = dedupe_rows(order_rows)
    summary_rows = dedupe_summary_rows(summary_rows)
    count = write_raw(out_path, order_rows)
    print(f"Wrote {count} Nextbite billing row(s) to {out_path}")
    if summary_out_path:
        summary_count = write_raw(summary_out_path, summary_rows)
        print(f"Wrote {summary_count} Nextbite summary row(s) to {summary_out_path}")
    return count


def main() -> None:
    parser = argparse.ArgumentParser(description="Extract Nextbite billings raw CSV from mbox.")
    parser.add_argument(
        "--mbox",
        default=takeout_path("Mail", "Billings-Nextbite.mbox"),
        help="Path to Billings-Nextbite.mbox",
    )
    parser.add_argument(
        "--out",
        default=raw_path("nextbite", "orders_raw.csv"),
        help="Output order-level raw CSV path.",
    )
    parser.add_argument(
        "--summary-out",
        default=raw_path("nextbite", "billings_raw.csv"),
        help="Output summary billings CSV path.",
    )
    args = parser.parse_args()
    run(args.mbox, args.out, args.summary_out)


if __name__ == "__main__":
    main()
